# evals/build_outreach_xlsx.py — build the outreach workbook from finished
# report JSONs (reports/<key>.json, written by run_deep_reports).
#
# One row per company: contact data (from the YC outreach sheet) + the report
# distilled into a SENDABLE email draft — kind, credit-first, grounded, with
# the full report link left as a placeholder to paste once the page is hosted.
# Usage: python -m evals.build_outreach_xlsx

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
REPORTS = ROOT / "reports"
OUT = ROOT / "outreach.xlsx"

# Contact rows from the outreach spreadsheet (gid=34506927).
CONTACTS = {
    "narrative": {
        "company": "Narrative",
        "founder": "Suchit Dubey",
        "title": "Chief Executive Officer",
        "site": "https://www.trynarrative.com/",
        "email": "suchit@trynarrative.com",
        "phone": "",
        "linkedin": "http://www.linkedin.com/in/dubeysuchit",
    },
    "vortexify": {
        "company": "VortexifyAI",
        "founder": "Aditya Tewari",
        "title": "Co-Founder",
        "site": "https://www.vortexify.ai",
        "email": "atewari@praxis-tech.ai",
        "phone": "+15138074955",
        "linkedin": "http://www.linkedin.com/in/adityatewari",
    },
}

HEADERS = [
    "Company", "Founder", "Title", "Site", "Founder Email", "Phone",
    "LinkedIn", "Report Status", "Score Hint", "Report Link",
    "Sendable Report (email draft)",
]

# Drive share links, filled AFTER you upload the PDFs to your own Drive.
# Edit web/dist/links.json:  {"vortexify": "https://drive.google.com/…", …}
# then re-run this builder — the link lands in the Report Link column AND
# replaces <REPORT LINK> inside each email draft.
LINKS_FILE = Path(__file__).resolve().parent.parent / "web" / "dist" / "links.json"


def _load_links() -> dict:
    try:
        return json.loads(LINKS_FILE.read_text(encoding="utf-8"))
    except (FileNotFoundError, json.JSONDecodeError):
        return {}


def _first_name(full: str) -> str:
    return full.split()[0] if full else "there"


def _obs_lines(observations: list[dict], limit: int) -> list[str]:
    out = []
    for o in observations[:limit]:
        claim = (o.get("claim") or "").strip().rstrip(".")
        if claim:
            out.append(f"  - {claim}.")
    return out


def _sugg_lines(suggestions: list[dict], limit: int) -> list[str]:
    out = []
    for s in suggestions[:limit]:
        sug = (s.get("suggestion") or s.get("claim") or "").strip().rstrip(".")
        if sug:
            out.append(f"  - {sug}.")
    return out


def sendable(contact: dict, rep: dict, pages: list[str], link: str = "") -> str:
    """Distill a report dict into a paste-ready outreach email."""
    fn = _first_name(contact["founder"])
    company = rep.get("company") or contact["company"]

    strengths = _obs_lines(rep.get("standout_strengths", []), 3)
    friction = _obs_lines(rep.get("friction_points", []), 2)
    improvements = _sugg_lines(rep.get("improvement_opportunities", []), 3)

    parts = [
        f"Subject: A first-impression read of {company}'s site — what a new visitor actually sees",
        "",
        f"Hi {fn},",
        "",
        f"I built First Impression Engine (FIE) — an agentic AI pipeline that reads a "
        f"company's public site the way a first-time visitor would, then reports back what "
        f"lands, what confuses, and what's missing. Every observation is pulled from a real "
        f"page and cited; nothing is invented. I ran it on {company} "
        f"({len(pages)} public pages) and thought you might find the outside view useful.",
        "",
        "What stood out (credit where due):",
        *(strengths or ["  - (see full report)"]),
    ]
    if friction:
        parts += ["", "Where a first-time visitor may hesitate:", *friction]
    if improvements:
        parts += ["", "Highest-leverage improvements the analysis surfaced:", *improvements]
    parts += [
        "",
        "The full report — persona-by-persona reads, every claim cited to the exact page — "
        f"is here: {link or '<REPORT LINK>'}",
        "",
        "This reflects only your public pages as crawled (robots.txt respected), so anything "
        "behind auth or added since isn't covered. If any point is off, that itself is signal "
        "about what the site communicates.",
        "",
        "Happy to share how it works under the hood.",
        "",
        "Keshav Kakani",
        "kkakani160@gmail.com · +91 90240 99116",
    ]
    return "\n".join(parts)


def main() -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Outreach"
    ws.append(HEADERS)
    for c in ws[1]:
        c.font = Font(bold=True)

    widths = [16, 18, 22, 34, 28, 15, 38, 14, 12, 40, 110]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    links = _load_links()
    missing_links = []
    for key, contact in CONTACTS.items():
        link = links.get(key, "")
        if not link:
            missing_links.append(key)
        path = REPORTS / f"{key}.json"
        if not path.exists():
            status, hint, draft = "NOT RUN", "", ""
        else:
            row = json.loads(path.read_text(encoding="utf-8"))
            if row.get("ok"):
                rep = row["report"]
                verdicts = row.get("panel_verdicts", {})
                yes = sum(1 for v in verdicts.values() if v)
                status = "READY" if link else "READY (add link)"
                hint = f"{yes}/{len(verdicts)} personas positive" if verdicts else ""
                draft = sendable(contact, rep, row.get("pages_examined", []), link)
            else:
                status = f"FAILED: {row.get('error_type')}"
                hint, draft = "", (row.get("error") or "")[:300]
        ws.append([
            contact["company"], contact["founder"], contact["title"], contact["site"],
            contact["email"], contact["phone"], contact["linkedin"], status, hint,
            link or "— upload PDF, add to links.json —", draft,
        ])

    # wrap the draft column, top-align everything
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 320
        for c in range(1, len(HEADERS) + 1):
            ws.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=(c == len(HEADERS)))

    wb.save(OUT)
    print(f"[xlsx] wrote {OUT}")
    if missing_links:
        print(f"[xlsx] NO Drive link yet for: {', '.join(missing_links)} — "
              f"upload the PDF(s), then add to {LINKS_FILE} and re-run.")


if __name__ == "__main__":
    main()
